# print("Varnitha Ediriweera")
# print('o----')
# print(' ||||')
# print('*'* 10)
from itertools import count
from time import sleep

from Tools.scripts.make_ctype import values
from unicodedata import digit

# price = 10
# price = 20
# rating = 4.9
# name = 'Varnitha'
# is_published = False
# print(price)

# your_name = input('What is your name? ')
# favorite_color = input('What your favorite color?')
# print('Hi ' + your_name)
# print(your_name + ' likes '+favorite_color)

# birth_year = input('Birth year: ')
# birth_year_int = int(birth_year)
# print(type(birth_year))
# print(type(birth_year_int))
# age = 2019 - birth_year_int
# print(age)

# weight_lbs = input('Weight (lbs): ')
# weight_kg = int(weight_lbs) * 0.45
# print(weight_kg)

# course = 'Python for "Beginners"'
# print(course)

# course = '''
#
# Hi Varnitha,
#
# Hera is our first email to you
#
# The support team.
#
# '''
# print(course)

# course = 'Python for Beginners'
# another = course[:]
# print(course[:5])

# first = 'John'
# last = 'Smith'
# message = first + ' ['+last+'] is a coder'
# msg = f'{first} [{last}] is a coder'
# print(message)
# print(msg)

# course = 'Python for Beginners'
# print(len(course))
# print(course.upper())
# print(course.find('f'))
# print(course.replace('Beginners','Absolute beginners'))
# print('Python' in course)

# print(10/3)
# print(10//3)
# print(10%3)
# print(10**3)

# x = 10
# x = x + 3
# x += 3
# x -= 1
# print(x)

# import math
#
# print(math.ceil(2.9))
# print(math.floor(2.9))
# x = 2.9
# print(round(x))
# print(abs(-2.9))

# is_hot = False
# is_cold = True
#
# if is_hot:
#     print("It's a hot day")
#     print("Drink plenty of water")
# elif is_cold:
#     print("It's a cold day")
#     print("Wear warm clothes")
# else:
#     print("It's a lovely day")
# print("Enjoy your day")

# has_high_income = True
# has_good_credit = True
#
# if has_high_income and has_good_credit:
#     print("Eligible for loan")

# has_good_credit = True
# has_criminal_record = False
#
# if has_good_credit and not has_criminal_record:
#     print("Eligible for loan")

# temperature = 35
#
# if temperature > 30:
#     print("It's a hot day")
# else:
#     print("It's not a hot day")
#
# guess_count = 3
# now_count = 0
# secret_number = 14
#
# while now_count < guess_count:
#     now_count += 1
#     entered_value = int(input('Guess : '))
#     if entered_value == secret_number:
#         if now_count == 1:
#             print("Your first guess is right!!!")
#         elif now_count == 2:
#             print("Your second guess is right!!!")
#         elif now_count == 3:
#             print("Your third guess is right!!!")
#         print("You Won")
#         break
#     else:
#         print("try again.")
# else:
#     print("Your try out times passed!")

# for item in ['Varnitha','Chamathkara','Ediriweera'] :
#     print(item)

# for item in range(5,10) :
#     print(item)

# for item in range(5, 10, 2):
#     print(item)

# for x in range(4):
#     for y in range(3):
#         print(f'({x}, {y})')

# number_list = [13, 24, 56, 89, 7, 1]
# max = number_list[0]
# min = number_list[0]
# for number in number_list:
#     if number > max:
#         max = number
#     else:
#         min = number
# print(f'Number List : {number_list}')
# print(f'MAX : {max}')
# print(f'MIN : {min}')

#

# numbers = [5, 2, 1, 7, 4]
# numbers2 = numbers.copy()
# numbers.insert(3,9)
# numbers.append(10)
# print(numbers)
# print(numbers2)

# numbers = [2, 4, 4, 5, 9, 5, 8]
# uniques = []
# for number in numbers:
#     if number not in uniques:
#         uniques.append(number)
# print(uniques)

# coordinates = (1, 2, 3)
# x, y, z = coordinates
# print(y)

# customer = {
#     "name": "Varnitha Chamathkara",
#     "age": 30,
#     "is_verified": True
# }
# print(customer["name"])
# customer["name"] = "Varnitha Ediriweera"
# print(customer.get("age"))
# print(customer["name"])

# phone = input("Phone No. : ")
# digit_mapping = {
#     "0": "Zero",
#     "1": "One",
#     "2": "Two",
#     "3": "Three",
#     "4": "Four",
#     "5": "Five",
#     "6": "Six",
#     "7": "Seven",
#     "8": "Eight",
#     "9": "Nine"
# }
#
# output = ""
# for phone_index in phone:
#     output += digit_mapping.get(phone_index,"!") + " "
#
# print(output)

# message = input(">")
# words = message.split(' ')
# emojis = {
#     ":)": "🙂",
#     ":(": "🙁"
# }
# output = ""
# for word in words:
#     output += emojis.get(word, word) + " "
# print(output)

# def greet_user():
#     print('Hi there!')
#     print('Welcome aboard')
#
# print("Start")
# greet_user()
# print("Finish")

# try:
#     age = int(input('Age: '))
#     print(age)
# except ValueError:
#     print('Invalid value')

# try:
#     age = int(input('Age: '))
#     income = 20000
#     risk = income / age
#     print(age)
# except ZeroDivisionError:
#     print('Age cannot be 0.')
# except ValueError:
#     print('Invalid value')

# class Point:
#     def __init__(self, x, y):
#         self.x = x
#         self.y = y
#
#     def move(self):
#         print("move")
#
#     def draw(self):
#         print("draw")


# point1 = Point()
# point1.x = 10
# point1.y = 20
# print(point1.x)
# point1.draw()

# point2 = Point()
# point2.x = 1
# print(point2.x)

# point = Point(10,20)
# print(point.x)

# class Person:
#
#     def __init__(self, name):
#         self.name = name
#
#     def talk(self):
#         print(f"Hi, I am {self.name}")
#
#
# john = Person("John Smith")
# john.talk()

# class Mammal:
#     def walk(self):
#         print("walk")
#
#
# class Dog(Mammal):
#     def bark(self):
#         print("bark")
#
#
# class Cat(Mammal):
#     def be_annoyingb(self):
#         print("annoying")
#
#
# dog1 = Dog()
# dog1.walk()
#
# cat1 = Cat()
# cat1.be_annoyingb()

# import converters
# from converters import kg_to_lbs
#
# print(converters.kg_to_lbs(70))
# print(kg_to_lbs(150))

# from utils import find_max
#
# print(find_max([10, 3, 16, 2]))

# from ecommerce import shipping
#
# shipping.calc_shipping()
# shipping.calc_shipping()
# shipping.calc_shipping()

# import random
#
# members = ['John', 'Mary', 'Bob', 'Mosh']
# leader = random.choice(members)
# print(leader)

# import random
#
# class Dice:
#     def roll(self):
#         first = random.randint(1, 6)
#         second = random.randint(1, 6)
#
#         return first, second
#
#
# dice = Dice()
# print(dice.roll())

# from pathlib import Path
#
# path = Path("emails")
# print(path.exists())
# print(path.mkdir())
# print(path.exists())
# print(path.rmdir())
# print(path.exists())

# from pathlib import Path
#
# path =  Path()
# for file in path.glob('*'):
#     print(file)

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # cell = sheet['a1']
    # cell = sheet.cell(1, 1)
    # print(sheet.max_row)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)
